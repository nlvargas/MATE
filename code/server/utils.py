from io import BytesIO
import gzip
import os
import json
import smtplib
from openpyxl import Workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


def student_to_student_type_id(student, hasModules):
    if hasModules:
        student_type_id = f"{student['attributes']} - {student['preferences']} - {student['disponibilities']}"
    else:
        student_type_id = f"{student['attributes']} - {student['preferences']}"
    return student_type_id


def get_report(results, students_preferences_number, hasModules):
    priority = results["priority"]
    groups = results["results"]
    x_axis = [i for i in range(1, students_preferences_number + 1)] + ['none']
    values = {x: 0 for x in x_axis}
    for group in groups:
        for student in group["students"]:
            student_type_id = student_to_student_type_id(student, hasModules)
            student_priority = priority[student_type_id][group["group"]]
            if student_priority not in (100, 1000):
                values[student_priority] += 1
            else:
                values["none"] += 1
    return values


def create_excel(results, modules, students_preferences_number, attributes, filepath):
    header = ["ID"] \
             +  attributes \
             + [f"Dispnibilidad {m}" for m in modules] \
             + [f"Preferencia {i}" for i in range(1, students_preferences_number + 1)]
    wb = Workbook()
    ws = wb.active
    ws.title = "Grupos"
    for i, g in enumerate(results["results"]):
        group_info = g["group_name"]
        group_name = f"Grupo {i+1}: {group_info}"
        ws.append([group_name])
        ws.append(header)
        for student in g["students"]:
            student["group_name"] = group_name
            ws.append([student["id"]] \
                       + [student["attributes"][a] for a in attributes] \
                       + student["disponibilities"] \
                       + [student["preferences"][str(i)] for i in range(1, students_preferences_number + 1)])
        ws.append([])
    
    report_ws = wb.create_sheet("Reporte")
    report = get_report(results, students_preferences_number, len(modules) != 0)
    total = sum(report[i] for i in report)
    for i in range(1, students_preferences_number + 1):
        report_ws.append([f"Cantidad de alumnos que quedaron en su preferencia {i}", f"{report[i]}/{total}"])
    report_ws.append([f"Cantidad de alumnos no quedaron en alguna de sus preferencias", f"{report['none']}/{total}"])

    students_ws = wb.create_sheet("Alumnos")
    students_ws.append(header + ["Grupo"])
    for i, g in enumerate(results["results"]):
        for student in g["students"]:
            students_ws.append([student["id"]] \
                                + [student["attributes"][a] for a in attributes] \
                                + student["disponibilities"] \
                                + [student["preferences"][str(i)] for i in range(1, students_preferences_number + 1)] \
                                + [student["group_name"]])

    wb.save(filepath)


def send_mail(receiver_address, filepath=None):
    sender_address = 'mate.ing.puc@gmail.com'
    sender_pass = 'mateingpuc123'
    message = MIMEMultipart()
    message['From'] = sender_address
    message['To'] = receiver_address
    message['Subject'] = 'Resultados MATE' 
    if filepath is None:
        mail_content = "No fue posible obtener una solución que cumpla con todas los requisitos. Por favor flexibiliza alguna restricción e inténtalo de nuevo."
        message.attach(MIMEText(mail_content, 'plain'))
    else:
        mail_content = "Adjuntamos los resultados obtenidos utilizando MATE"
        message.attach(MIMEText(mail_content, 'plain'))
        part = MIMEBase('application', "octet-stream")
        part.set_payload(open(filepath, "rb").read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="results.xlsx"')
        message.attach(part)
    session = smtplib.SMTP('smtp.gmail.com', 587) 
    session.starttls()
    session.login(sender_address, sender_pass)
    text = message.as_string()
    session.sendmail(sender_address, receiver_address, text)
    session.quit()


def send_admin_mail(receiver_address, xlsx_filepath, txt_filepath):
    sender_address = 'mate.ing.puc@gmail.com'
    sender_pass = 'mateingpuc123'
    message = MIMEMultipart()
    message['From'] = sender_address
    message['To'] = receiver_address
    message['Subject'] = 'Resultados MATE (Admin)' 
    mail_content = "Adjuntamos los resultados obtenidos utilizando MATE"
    message.attach(MIMEText(mail_content, 'plain'))

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(xlsx_filepath, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="results.xlsx"')
    message.attach(part)

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(txt_filepath, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="result.txt"')
    message.attach(part)

    session = smtplib.SMTP('smtp.gmail.com', 587) 
    session.starttls()
    session.login(sender_address, sender_pass)
    text = message.as_string()
    session.sendmail(sender_address, receiver_address, text)
    session.quit()


def send_error_mail(receiver_address):
    sender_address = 'mate.ing.puc@gmail.com'
    sender_pass = 'mateingpuc123'
    message = MIMEMultipart()
    message['From'] = sender_address
    message['To'] = receiver_address
    message['Subject'] = 'Resultados MATE' 
    mail_content = "Ha ocurrido un error al obtener la asignación de grupos. Por favor inténtelo de nuevo."
    message.attach(MIMEText(mail_content, 'plain'))
    session = smtplib.SMTP('smtp.gmail.com', 587) 
    session.starttls()
    session.login(sender_address, sender_pass)
    text = message.as_string()
    session.sendmail(sender_address, receiver_address, text)
    session.quit()


def print_data(ID, data):
    print("ID:", ID)
    print("Email:", json.dumps(data["email"], indent=4, default=str))
    print("Students number:", len(data["students"]))
    # print("Students types:", json.dumps(data["students_types"], indent=4, default=str))
    print("Groups number:", json.dumps(data["groups_number"], indent=4, default=str))
    print("Min. students number:", json.dumps(data["lower_number"], indent=4, default=str))
    print("Max. students number:", json.dumps(data["upper_number"], indent=4, default=str))
    print("Same day:", json.dumps(data["sameDay"], indent=4, default=str))
    print("Attributes:", json.dumps(data["attributes"], indent=4, default=str))
    print("Options:", json.dumps(data["A"], indent=4, default=str))
    print("Fixed day:", json.dumps(data["fixedDay"], indent=4, default=str))
    print("Preferences:", json.dumps(data["preferences"], indent=4, default=str))
    print("Preferences number:", json.dumps(data["usedPreferences"], indent=4, default=str))
    print("Modules:", json.dumps(data["modules"], indent=4, default=str))
    print("Capacity:", json.dumps(data["capacity"], indent=4, default=str))
    print("\n")
    

def get_data(ID):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    compressed_filename = f"{BASE_DIR}/groups/params/{ID}_params.txt"
    compressed_data = open(compressed_filename, "rb")
    compressed_data_string_bytes = compressed_data.read()
    data_string = decompress_bytes_to_string(compressed_data_string_bytes)
    compressed_data.close()
    data = json.loads(data_string)
    attr_bounds = data["attributes"]
    for a in data["attributes"]:
        for r in data["attributes"][a]:
            if data["attributes"][a][r]["min"] is None:
                data["attributes"][a][r]["min"] = 0
            if data["attributes"][a][r]["max"] is None:
                data["attributes"][a][r]["max"] = int(data["upper_number"])
    for p in data["preferences"]:
        if data["preferences"][p]["min"] is None:
            data["preferences"][p]["min"] = 0
        if data["preferences"][p]["max"] is None:
            data["preferences"][p]["max"] = int(data["groups_number"])
    
    print_data(ID, data)
    
    return data


def decompress_bytes_to_string(inputBytes):
  """
  decompress the given byte array (which must be valid 
  compressed gzip data) and return the decoded text (utf-8).
  """
  bio = BytesIO()
  stream = BytesIO(inputBytes)
  decompressor = gzip.GzipFile(fileobj=stream, mode='r')
  while True:  # until EOF
    chunk = decompressor.read(8192)
    if not chunk:
      decompressor.close()
      bio.seek(0)
      return bio.read().decode("utf-8")
    bio.write(chunk)
  return None